# 3-1. Create Excel Workbook
from openpyxl import Workbook

filename = '사원_정보 2.xlsx'
# Workbook 객체 생성
e_info_wb = Workbook()
# 활성화된 워크시트 변수 할당
active_ws = e_info_wb.active
# 제목 설정
active_ws.title = '인사_기록'
# 워크북에 '과거_인사_평가', '연봉' 워크시트 추가
e_info_wb.create_sheet('과거_인사_평가')
e_info_wb.create_sheet('연봉')
# 워크북 저장
e_info_wb.save(filename)

# 3-2. 엑셀 파일 불러오기
from openpyxl import load_workbook

filename = '사원_정보 2.xlsx'
# load excel
e_info_wb = load_workbook(filename)
# 워크시트 목록 출력
print(e_info_wb.sheetnames)
# '인사_기록' 워크시트 선택
personnel_info_ws = e_info_wb['인사_기록']
# 선택된 워크시트 정보 출력
print(personnel_info_ws)
# 워크시트 제거
import openpyxl
e_info_wb2 = openpyxl.load_workbook(filename)
e_info_wb2.sheetnames
remove_ws = e_info_wb2.get_sheet_by_name('인사_기록')
e_info_wb2.remove_sheet(remove_ws)
e_info_wb2.sheetnames

#3-3. 칼럼과 인덱스 간의 변환
from openpyxl.utils import get_column_letter, column_index_from_string
# 열 'A'에 대한 인덱스 값
print(column_index_from_string('A'))
# 열 'Z'에 대한 인덱스 값
print(column_index_from_string('Z'))
# 열 'ZZ'에 대한 인덱스 값
print(column_index_from_string('ZZ'))
# 인덱스 1 에 대한 열 문자 값
print(get_column_letter(1))
# 인덱스 26 에 대한 열 문자 값
print(get_column_letter(26))
# 인덱스 702 에 대한 열 문자 값
print(get_column_letter(702))

#3-4. 리스트형 자료를 열 단위로 이어 붙이기
from operator import itemgetter
from openpyxl import load_workbook
filename = '사원_정보 2.xlsx'
# load excel
e_info_wb = load_workbook(filename)
# 워크시트 목록 출력
print(e_info_wb.sheetnames)
# '인사_기록' 워크시트 선택
personnel_info_ws = e_info_wb['인사_기록']
# 추가될 컬럼의 목록
column = ['ID', '이름', '주소', '생년월일', '학력', '입사 연월']
# append method 사용
personnel_info_ws.append(column)
# 추가될 데이터
data = [['1233', '박찬성', '세종시 도담동', '1985-12-23', '학사', '2012-11'],
 ['1723', '김진수', '대전시 월평동', '1989-05-17', '학사', '2017-02'],
 ['1435', '이수빈', '세종시 아름동', '1988-02-25', '박사', '2014-05'],
 ['1601', '김민준', '대전시 반석동', '1989-07-27', '석사', '2016-12'],
 ['1804', '최형아', '대전시 둔산동', '1986-01-20', '학사', '2018-03']]

# 첫번째 키 기준 정렬
# key: Optional. A Function to execute to decide the order.
# Default is None
# itemgetter(0) : index 0 을 기준으로 sort 하겠다는 의미
data = sorted(data, key=itemgetter(0))
data

# 이름 순으로 정렬
data = sorted(data, key=itemgetter(1))
data
# for 문과 append 메서드 사용
for row in data:
 personnel_info_ws.append(row)
filename2 = '사원_정보 3.xlsx'
e_info_wb.save(filename2)

#3-5. 행 단위로 각 열에 맞는 데이터 추가 삽입
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
filename = '사원_정보 3.xlsx'
# load excel
e_info_wb = load_workbook(filename)
# '인사_기록' 워크시트 선택
personnel_info_ws = e_info_wb['인사_기록']
# 행 최대 크기와 해당 행의 알파벳
print(personnel_info_ws.max_column)
print(get_column_letter(personnel_info_ws.max_column))
# 5 번째 행에 하나의 행 추가
# insert_cols(idx, amount=1)
# idx 에 해당하는 열 바로 앞에 col 추가
personnel_info_ws.insert_cols(5,1)
# personnel_info_ws.insert_cols(5,2) # 열 2 개
# 행 추가 후 행 최대 크기 조사
print(personnel_info_ws.max_column)
print(get_column_letter(personnel_info_ws.max_column))
# 추가된 행에 추가할 데이터
new_column = '병역'
new_data = ['군필', '미필', '제외', '군필', '제외']
# E1 셀에 '병역' 문자열 추가
personnel_info_ws['E1'] = new_column
# 첫 번째 행은 열 이름으로 사용되고 두번째 행부터 데이터 삽입을 위한 카운터
count = 2
# 추가된 행의 열마다 해당 데이터를 넣어줌
for data in new_data:
 personnel_info_ws.cell(row = count, column = 5, value=data)
 count = count +1 # count +=
e_info_wb
filename3 = '사원_정보 4.xlsx'
e_info_wb.save(filename3)

# numpy & pandas 이용
import pandas as pd
import numpy as np
# df = pd.read_excel('사원정보 4.xlsx')
df = pd.read_excel(filename3)
# df.insert(3, "column1", np.nan)
df
new_column = '재직여부'
add_data = ['재직', '이직', '재직', '재직', '휴직']
df.insert(3, new_column, add_data)
df
filename4 = '사원_정보 5.xlsx'
df.to_excel(filename4)

# 여러 worksheet 를 각각 dataframe 으로 만들어 excel 로 저장
import pandas as pd
import numpy as np
import xlsxwriter
ws1_contents = {
 'name1': ['김', '이', '박', '최'],
 'age': [10, 20, 30, 40]
}
ws2_contents = {
 'name2': ['정', '강', '조', '윤'],
 'age': [11, 21, 31, 41]
}
ws3_contents = {
 'name3': ['장', '임', '한', '오'],
 'age': [12, 22, 32, 42]
}
# Dictionary to Dataframe conversion
dataframe1 = pd.DataFrame(ws1_contents)
dataframe2 = pd.DataFrame(ws2_contents)
dataframe3 = pd.DataFrame(ws3_contents)
with pd.ExcelWriter('employee1.xlsx', engine='xlsxwriter') as writer:
 dataframe1.to_excel(writer, sheet_name='worksheet1')
 dataframe2.to_excel(writer, sheet_name='worksheet2')
 dataframe3.to_excel(writer, sheet_name='worksheet3')

#3-7. 규칙파일을 읽어서 해당 내용을 반환
#3-8. get_rules 함수 실행
def get_rules(filename):
 file = open(filename, 'r')
 key_column = file.readline().split(':')[1] # 1 로 수정
 key = file.readline().split(':')[1]
 targets = []
 for tmp_line in file:
 line = tmp_line.strip()
 wb_name = line.split(':')[0]
 ws_names = line.split(':')[1].split(',')
 target = {wb_name : ws_names}
 targets.append(target)
 return key_column.strip(), key.strip(), targets

# 코드 3-8 get_rules 함수의 시험
key_column, key, targets = get_rules('규칙파일.txt')
print('참조 키: ' + key_column)
print('참조 키 값: ' + key)
print('대상: ' + str(targets))

#3-9. 워크북에서 원하는 열을 찾아서 반환
from openpyxl import load_workbook
from openpyxl import Workbook
# 참조 키로 검색 대상이 될 열을 찾는 함수
def find_column(ws, key_column):
  # 열을 찾지 못하면 -1
  column = -1
  # 엑셀의 셀 번호는 1 부터 시작
  for col in range(1, ws.max_column + 1):
    if ws.cell(row=1, column=col).value == key_column:
        # print(ws.cell(row=1, column=col).value)
        column = col
        break
  return column


# 참조 키 값으로 검색 대상 행 목록을 가져오는 함수
# 각 행은 Cell 이라는 객체의 목록으로 구성된다.
def get_rows(ws, key, column):
 rows = []
 for row in range(1, ws.max_row + 1):
  if ws.cell(row=row, column=column).value == int(key):
   rows.append(ws[row])

 return rows


# Cell 객체로부터 값을 추출하는 기능의 함수
def get_raw_row(row):
 cells = []
 for cell in row:
  cells.append(cell.value)

 return cells


# 검색된 열의 내용을 배열에 저장 후 반환해 주는 함수
def get_row_by(key_column, key, target):
   # 키만 모아서 리스트로 변환
   wb_name = list(target.keys())[0]
   #모든 워크시트 목록
   ws_names = target[wb_name]
   rows = []
   wb = load_workbook(wb_name)
   for ws_name in ws_names:
       ws = wb[ws_name]
       column = find_column(ws, key_column)
       # 참조 키가 발견된 경우에만 수행
       if column != -1:
          rows_cell = get_rows(ws, key, column)
          # 행 목록에서 각 행을 for 반복문으로 접근
          for row in rows_cell:
          # 문자열 목록으로 변환 후 행 추가
               rows_raw = get_raw_row(row)
               rows.append(rows_raw)
   return rows

#3-10. 워크북에서 원하는 열을 찾아서 반환
from openpyxl import load_workbook
from openpyxl import Workbook
def get_rules(filename):
   file = open(filename, 'r')
   key_column = file.readline().split(':')[1]
   key = file.readline().split(':')[1]
   targets = []
   for tmp_line in file:
      line = tmp_line.strip()
      wb_name = line.split(':')[0]
      ws_names = line.split(':')[1].split(',')
      target = {wb_name : ws_names}
      targets.append(target)
   return key_column.strip(), key.strip(), targets

def find_column(ws, key_column):
   column = -1
   for col in range(1, ws.max_column+1):
      if ws.cell(row=1, column=col).value == key_column:
            # print(ws.cell(row=1, column=col).value)
            column = col
            break
   return column

def get_rows(ws, key, column):
     rows = []
     for row in range(1, ws.max_row+1):
         if ws.cell(row=row, column=column).value == int(key):
                rows.append(ws[row])
     return rows

def get_raw_row(row):
    cells = []
    for cell in row:
         cells.append(cell.value)
    return cells

def get_row_by(key_column, key, target):
    wb_name = list(target.keys())[0]
    ws_names = target[wb_name]
    rows = []
    wb = load_workbook(wb_name)
    for ws_name in ws_names:
          ws = wb[ws_name]
          column = find_column(ws, key_column)
          if column != -1:
               rows_cell = get_rows(ws, key, column)
               for row in rows_cell:
                    rows_raw = get_raw_row(row)
                    rows.append(rows_raw)
    return rows


def main():
      key_column, key, targets = get_rules('규칙파일.txt')
      rows = []
      merge_info_wb = Workbook()
      active_ws = merge_info_wb.active
      active_ws_title = '취합_' + key
      for target in targets:
           tmp_rows = get_row_by(key_column, key, target)
           for row in tmp_rows:
                     active_ws.append(row)
      merge_info_wb.save('취합_정보_' + key + '_2' + '.xlsx')

# if __name__ == "__main__":
# main()

